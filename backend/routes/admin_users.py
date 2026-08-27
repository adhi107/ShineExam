# Shine Exam admin routes for candidate account management.
import io
import csv
import re
from datetime import datetime, timedelta, time
from bson import ObjectId
from typing import Optional
from flask import Blueprint, jsonify, request, send_file
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from config.db import get_db
from utils.json import to_jsonable
from utils.validators import require_fields

admin_users_bp = Blueprint("admin_users", __name__)

EMAIL_REGEX = re.compile(r"^[^\s@]+@[^\s@]+\.[^\s@]+$")


def _parse_valid_until(value):
    if value in (None, ""):
        return None
    try:
        parsed = datetime.fromisoformat(str(value).strip().replace("Z", "+00:00"))
        if len(str(value).strip()) == 10:
            parsed = datetime.combine(parsed.date(), time.max)
        return parsed.replace(tzinfo=None) if parsed.tzinfo else parsed
    except (TypeError, ValueError):
        raise ValueError("Validity date must use YYYY-MM-DD format")


def _merge_registration_fields(user: dict, registration: Optional[dict]) -> dict:
    if not registration:
        return user

    merged = dict(user)
    field_map = {
        "naxUnid": "naxUnid",
        "studentName": "name",
        "studentId": "studentId",
        "email": "email",
        "collegeEmail": "collegeEmail",
        "mobile": "mobile",
        "gender": "gender",
        "courseStream": "courseStream",
        "cgpa": "cgpa",
        "sapCertification": "sapCertification",
        "collegeName": "collegeName",
    }

    for reg_key, user_key in field_map.items():
        if merged.get(user_key) in (None, ""):
            merged[user_key] = registration.get(reg_key)

    if merged.get("collegeRollNumber") in (None, "") and registration.get("studentId"):
        merged["collegeRollNumber"] = registration.get("studentId")

    return merged

# List candidate accounts and expire access when validity dates pass.
@admin_users_bp.route("", methods=["GET"])
@admin_users_bp.route("/", methods=["GET"])
def list_users():
    db = get_db()
    now = datetime.utcnow()
    db.users.update_many(
        {"role": "answerer", "validUntil": {"$lt": now}, "isActive": {"$ne": False}},
        {"$set": {"isActive": False, "statusReason": "validity_expired", "statusUpdatedAt": now}},
    )
    users = list(db.users.find({"role": "answerer"}, {"password": 0}))
    out = []
    for u in users:
        out.append({
            "id": str(u["_id"]), "name": u.get("name") or u.get("userId"),
            "email": u.get("email", ""), "userId": u.get("userId"),
            "createdAt": u.get("createdAt"), "lastLoginAt": u.get("lastLoginAt"),
            "isActive": u.get("isActive", True),
            "statusReason": u.get("statusReason", ""),
            "statusUpdatedAt": u.get("statusUpdatedAt"),
            "blockedDueTo": u.get("blockedDueTo", ""),
            "validUntil": u.get("validUntil"),
            "isExpired": bool(u.get("validUntil") and u.get("validUntil") < now),
            "attempts": db.results.count_documents({"userId": u.get("userId")}),
        })
    return jsonify({"users": to_jsonable(out)})


# Create a candidate account from the admin student directory.
@admin_users_bp.route("", methods=["POST"])
@admin_users_bp.route("/", methods=["POST"])
def create_user():
    payload = request.get_json(silent=True) or {}

    ok, msg = require_fields(payload, ["name", "email", "userId", "password"])
    if not ok:
        return jsonify({"error": msg}), 400
    
    db = get_db()
    userId = str(payload["userId"]).strip()
    if db.users.find_one({"userId": userId}):
        return jsonify({"error": "userId already exists"}), 409
    
    try:
        valid_until = _parse_valid_until(payload.get("validUntil"))
    except ValueError as exc:
        return jsonify({"error": str(exc)}), 400
    doc = {
        "name": payload["name"].strip(),
        "email": payload["email"].strip().lower(),
        "userId": userId,
        "password": str(payload["password"]).strip(),  # Stored to match the current Shine Exam login model.
        "role": "answerer",
        "createdAt": datetime.utcnow(),
        "lastLoginAt": None,
        "isActive": True,
        "validUntil": valid_until,
    }
    res = db.users.insert_one(doc)
    return jsonify({
        "user": to_jsonable({
            "id": str(res.inserted_id),
            "name": doc["name"],
            "email": doc["email"],
            "userId": doc["userId"],
            "role": doc["role"],
            "createdAt": doc["createdAt"],
            "isActive": doc["isActive"], 
            "validUntil": doc["validUntil"],
        })
    }), 201

# Let an administrator reset a candidate's Shine Exam password.
@admin_users_bp.route("/<user_id>/change-password", methods=["PUT", "PATCH"])
@admin_users_bp.route("/<user_id>/change-password/", methods=["PUT", "PATCH"])
def admin_change_user_password(user_id: str):
    """
    Admin endpoint to change any user's password.
    Payload:
    {
      "newPassword": "..."
    }
    """
    payload = request.get_json(silent=True) or {}
    ok, msg = require_fields(payload, ["newPassword"])
    if not ok:
        return jsonify({"error": msg}), 400
    
    new_password = str(payload["newPassword"]).strip()
    
    if not new_password or len(new_password) < 4:
        return jsonify({"error": "Password must be at least 4 characters"}), 400
    
    db = get_db()
    
    # Support resets from either the Mongo record id or the candidate username.
    try:
        user = db.users.find_one({"_id": ObjectId(user_id)})
    except Exception:
        user = db.users.find_one({"userId": user_id})
    
    if not user:
        return jsonify({"error": "User not found"}), 404
    
    # Save the new candidate password and timestamp the admin reset.
    db.users.update_one(
        {"_id": user["_id"]},
        {
            "$set": {
                "password": new_password,
                "passwordUpdatedAt": datetime.utcnow()
            }
        }
    )
    
    return jsonify({
        "message": "Password updated successfully",
        "userId": user.get("userId")
    })


# Block or unblock a candidate's portal access.
@admin_users_bp.route("/<user_id>/status", methods=["PUT", "PATCH"])
@admin_users_bp.route("/<user_id>/status/", methods=["PUT", "PATCH"])
def update_user_status(user_id: str):
    """
    Payload:
    {
        "isActive": true/false
    }
    """
    payload = request.get_json(silent=True) or {}

    if "isActive" not in payload:
        return jsonify({"error": "isActive field required"}), 400

    db = get_db()

    # Support status changes from either the Mongo record id or the candidate username.
    try:
        q = {"_id": ObjectId(user_id)}
    except Exception:
        q = {"userId": user_id}

    user = db.users.find_one(q)
    if not user:
        return jsonify({"error": "User not found"}), 404

    status_updates = {
        "isActive": bool(payload["isActive"]),
        "statusUpdatedAt": datetime.utcnow()
    }
    if status_updates["isActive"]:
        status_updates["statusReason"] = "unblocked_by_admin"
        status_updates["blockedDueTo"] = None
        if user.get("validUntil") and user.get("validUntil") < datetime.utcnow():
            status_updates["validUntil"] = None
    else:
        status_updates["statusReason"] = "manually_blocked"

    db.users.update_one(
        {"_id": user["_id"]},
        {
            "$set": status_updates
        }
    )

    return jsonify({
        "message": "User status updated",
        "userId": user.get("userId"),
        "isActive": bool(payload["isActive"])
    })


# Delete a candidate account and remove related test activity links.
@admin_users_bp.route("/<user_id>", methods=["DELETE"])
@admin_users_bp.route("/<user_id>/", methods=["DELETE"])
def delete_user(user_id: str):
    db = get_db()
    try:
        q = {"_id": ObjectId(user_id)}
    except Exception:
        q = {"userId": user_id}
    
    user = db.users.find_one(q)
    if not user:
        return jsonify({"error": "User not found"}), 404
    
    db.users.delete_one({"_id": user["_id"]})
    db.exam_assignments.delete_many({"userId": user.get("userId")})
    db.attempts.delete_many({"userId": user.get("userId")})
    
    return jsonify({"message": "Deleted"})


@admin_users_bp.route("/<user_id>", methods=["PUT", "PATCH"])
@admin_users_bp.route("/<user_id>/", methods=["PUT", "PATCH"])
def update_user(user_id: str):
    payload = request.get_json(silent=True) or {}
    db = get_db()
    try:
        q = {"_id": ObjectId(user_id)}
    except Exception:
        q = {"userId": user_id}

    user = db.users.find_one(q)
    if not user:
        return jsonify({"error": "User not found"}), 404

    allowed = ["name", "email"]
    updates = {k: payload[k] for k in allowed if k in payload}
    if "validUntil" in payload:
        try:
            updates["validUntil"] = _parse_valid_until(payload.get("validUntil"))
        except ValueError as exc:
            return jsonify({"error": str(exc)}), 400
    updates["updatedAt"] = datetime.utcnow()

    db.users.update_one({"_id": user["_id"]}, {"$set": updates})

    updated = db.users.find_one({"_id": user["_id"]}, {"password": 0})
    return jsonify({"user": to_jsonable({
        "id": str(updated["_id"]),
        **{k: updated.get(k) for k in ["name", "email", "userId", "createdAt", "lastLoginAt", "isActive", "validUntil"]}
    })})


# Download Excel/CSV template for student creation.
@admin_users_bp.route("/template", methods=["GET"])
@admin_users_bp.route("/template/", methods=["GET"])
@admin_users_bp.route("/excel-template", methods=["GET"])
def download_student_template():
    format_type = request.args.get("format", "xlsx").lower()
    
    headers = ["Full Name", "Email Address", "Username", "Password", "Valid Until (YYYY-MM-DD)"]
    default_validity = (datetime.utcnow() + timedelta(days=365)).strftime("%Y-%m-%d")
    
    sample_rows = [
        ["Aarav Patel", "aarav.patel@example.com", "STU202601", "Shine@2026", default_validity],
        ["Priya Sharma", "priya.sharma@example.com", "STU202602", "Shine@2026", default_validity],
    ]
    
    if format_type == "csv":
        output = io.StringIO()
        writer = csv.writer(output)
        writer.writerow(headers)
        for row in sample_rows:
            writer.writerow(row)
        mem = io.BytesIO()
        mem.write(output.getvalue().encode("utf-8"))
        mem.seek(0)
        return send_file(
            mem,
            mimetype="text/csv",
            as_attachment=True,
            download_name="student_creation_template.csv"
        )
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Student Accounts"
    
    header_fill = PatternFill(start_color="1E3A8A", end_color="1E3A8A", fill_type="solid")
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    align_center = Alignment(horizontal="center", vertical="center")
    align_left = Alignment(horizontal="left", vertical="center")
    thin_border = Border(
        left=Side(style="thin", color="E2E8F0"),
        right=Side(style="thin", color="E2E8F0"),
        top=Side(style="thin", color="E2E8F0"),
        bottom=Side(style="thin", color="E2E8F0")
    )
    
    ws.append(headers)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = align_center
    
    for row in sample_rows:
        ws.append(row)
        
    for row in ws.iter_rows(min_row=2, max_row=len(sample_rows)+1, min_col=1, max_col=5):
        for cell in row:
            cell.alignment = align_left
            cell.border = thin_border
            
    for col in ws.columns:
        max_len = max(len(str(cell.value or "")) for cell in col)
        col_letter = get_column_letter(col[0].column)
        ws.column_dimensions[col_letter].width = max(max_len + 5, 20)
        
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    return send_file(
        output,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name="student_creation_template.xlsx"
    )


def _parse_spreadsheet(file_obj, filename: str):
    rows = []
    ext = filename.lower().split(".")[-1] if "." in filename else ""
    
    if ext in ["xlsx", "xls"]:
        wb = openpyxl.load_workbook(file_obj, data_only=True)
        sheet = wb.active
        all_rows = list(sheet.iter_rows(values_only=True))
        if not all_rows:
            return []
        
        headers = [str(h).strip() if h is not None else "" for h in all_rows[0]]
        for r in all_rows[1:]:
            if not any(r):
                continue
            row_dict = {}
            for col_idx, header in enumerate(headers):
                if header and col_idx < len(r):
                    val = r[col_idx]
                    row_dict[header] = str(val).strip() if val is not None else ""
            rows.append(row_dict)
    else:
        content = file_obj.read()
        try:
            decoded = content.decode("utf-8-sig")
        except UnicodeDecodeError:
            decoded = content.decode("latin-1")
            
        reader = csv.DictReader(io.StringIO(decoded))
        for r in reader:
            row_dict = {k.strip(): str(v).strip() if v is not None else "" for k, v in r.items() if k}
            if any(row_dict.values()):
                rows.append(row_dict)
                
    return rows


# Bulk student account creation from uploaded Excel/CSV file.
@admin_users_bp.route("/bulk-upload", methods=["POST"])
@admin_users_bp.route("/bulk-upload/", methods=["POST"])
def bulk_upload_users():
    if "file" not in request.files:
        return jsonify({"error": "No file uploaded"}), 400
    
    file = request.files["file"]
    if not file or not file.filename:
        return jsonify({"error": "Selected file is empty"}), 400
    
    filename = file.filename
    try:
        raw_rows = _parse_spreadsheet(file.stream, filename)
    except Exception as exc:
        return jsonify({"error": f"Failed to parse file: {str(exc)}"}), 400
        
    if not raw_rows:
        return jsonify({"error": "The uploaded spreadsheet contains no data rows"}), 400
        
    db = get_db()
    now = datetime.utcnow()
    default_validity = datetime.combine((now + timedelta(days=365)).date(), time.max)
    
    created_users = []
    errors = []
    seen_usernames = set()
    
    existing_user_ids = set(doc["userId"] for doc in db.users.find({}, {"userId": 1}))
    
    for idx, row in enumerate(raw_rows, start=2):
        name = ""
        email = ""
        user_id = ""
        password = ""
        valid_until_str = ""
        
        for k, v in row.items():
            key_lower = k.lower().replace("_", " ").replace("-", " ")
            if key_lower in ["full name", "name", "student name", "candidate name"]:
                name = v
            elif key_lower in ["email address", "email", "student email", "college email"]:
                email = v
            elif key_lower in ["username", "user id", "userid", "student id", "roll number", "college roll number"]:
                user_id = v
            elif key_lower in ["password", "temp password", "temporary password"]:
                password = v
            elif key_lower in ["valid until (yyyy-mm-dd)", "valid until", "valid until date", "validity", "expiry date"]:
                valid_until_str = v

        row_errors = []
        if not name:
            row_errors.append("Full Name is required")
        if not email:
            row_errors.append("Email address is required")
        elif not EMAIL_REGEX.match(email):
            row_errors.append(f"Invalid email format '{email}'")
        if not user_id:
            row_errors.append("Username / Student ID is required")
        elif user_id in seen_usernames:
            row_errors.append(f"Duplicate Username '{user_id}' in this spreadsheet")
        elif user_id in existing_user_ids:
            row_errors.append(f"Username '{user_id}' already exists in system database")
            
        valid_until = default_validity
        if valid_until_str:
            try:
                valid_until = _parse_valid_until(valid_until_str)
            except ValueError as val_err:
                row_errors.append(str(val_err))

        if row_errors:
            errors.append({
                "row": idx,
                "name": name or "—",
                "userId": user_id or "—",
                "email": email or "—",
                "reason": "; ".join(row_errors)
            })
            continue

        seen_usernames.add(user_id)
        pwd = password if password else "Shine@2026"
        
        user_doc = {
            "name": name.strip(),
            "email": email.strip().lower(),
            "userId": user_id.strip(),
            "password": str(pwd).strip(),
            "role": "answerer",
            "createdAt": now,
            "lastLoginAt": None,
            "isActive": True,
            "validUntil": valid_until,
        }
        
        res = db.users.insert_one(user_doc)
        created_users.append({
            "id": str(res.inserted_id),
            "name": user_doc["name"],
            "email": user_doc["email"],
            "userId": user_doc["userId"],
            "role": user_doc["role"],
            "createdAt": user_doc["createdAt"],
            "isActive": user_doc["isActive"],
            "validUntil": user_doc["validUntil"]
        })
        
    return jsonify({
        "success": True,
        "totalRows": len(raw_rows),
        "createdCount": len(created_users),
        "failedCount": len(errors),
        "createdUsers": to_jsonable(created_users),
        "errors": errors
    }), 200
